"""Generate the synthetic Toast-shaped test fixtures.

Run ``python tests/fixtures/build_fixtures.py`` to regenerate every file in this
directory. The output is deterministic.

EVERY FILE THIS SCRIPT WRITES IS SYNTHETIC. See README.md in this directory.
"""

from __future__ import annotations

import csv
import random
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook

HERE = Path(__file__).resolve().parent

# --------------------------------------------------------------------------
# PMIX workbook 1: the "full" shape.
# Sheets: Summary, All levels, Items, Open items, Modifiers (empty),
# Special requests. `All levels` carries Type and mixes rollups with leaves.
# --------------------------------------------------------------------------
PMIX_FULL_ALL_LEVELS = [
    ["Type", "Menu", "Menu group", "Item, open item", "Qty sold"],
    # Rollup rows: blank Type. Counting these double-counts the whole file.
    ["", "LIQUOR", "", "", 400],
    ["", "LIQUOR", "Tequila", "", 193],
    ["menuItem", "LIQUOR", "Tequila", "Espolon Blanco", 193],
    ["", "LIQUOR", "Rum", "", 4],
    # Genuine liquor SKU with a BLANK Sales Category (no category column here at
    # all on this sheet -- the category lives on the Items sheet).
    ["menuItem", "LIQUOR", "Rum", "Bacardi FB", 4],
    ["", "COCKTAIL", "", "", 34],
    ["menuItem", "COCKTAIL", "House Cocktails", "PassionPunch Margarita", 34],
    ["", "BOTTLED BEER", "", "", 210],
    ["menuItem", "BOTTLED BEER", "Domestics", "Bud Light", 187],
    ["menuItem", "BOTTLED BEER", "Seltzer", "Nutrl", 23],
    ["", "Open items", "", "", 3],
    ["openItem", "Open items", "Open Drink", "Lobos Blanco Bottle", 3],
    # Same item name, different menu/category: bottle service, not a pour.
    ["menuItem", "BOTTLE SERVICE", "Tequila Bottles", "Espolon Blanco", 2],
    ["Total", "", "", "", 446],
]

PMIX_FULL_ITEMS = [
    ["Item", "Sales Category", "Qty sold"],
    ["Espolon Blanco", "Liquor", 195],
    ["Bacardi FB", "", 4],          # blank category on a real liquor SKU
    ["PassionPunch Margarita", "Liquor", 34],
    ["Bud Light", "Bottled Beer", 187],
    ["Nutrl", "Bottled Beer", 23],
    ["Red Bull", "NA Beverage", 41],
]

PMIX_FULL_OPEN_ITEMS = [
    ["Item", "Sales Category", "Qty sold"],
    ["Lobos Blanco Bottle", "Bottle Service", 3],
]

# --------------------------------------------------------------------------
# PMIX workbook 2: the variant shape from the same account.
# NO Sales Category column anywhere. HAS Subgroup, Avg. price, Gross/Net sales.
# Column ORDER differs, and the Modifiers / Special requests sheets are absent.
# --------------------------------------------------------------------------
PMIX_VARIANT_ALL_LEVELS = [
    ["Menu", "Menu group", "Subgroup", "Type", "Item, open item", "Avg. price", "Gross sales", "Net sales", "Qty sold"],
    ["LIQUOR", "Tequila", "Blanco", "", "", 9.0, 1737.0, 1700.0, 193],
    ["LIQUOR", "Tequila", "Blanco", "menuItem", "Espolon Blanco", 9.0, 1737.0, 1700.0, 193],
    ["LIQUOR", "Rum", "White", "menuItem", "Bacardi FB", 8.0, 32.0, 32.0, 4],
    ["BOTTLED BEER", "Domestics", "", "menuItem", "Bud Light", 5.0, 935.0, 900.0, 187],
    ["Open items", "Open Drink", "", "openItem", "Lobos Blanco Bottle", 220.0, 660.0, 660.0, 3],
]

PMIX_VARIANT_ITEMS = [
    ["Qty sold", "Item"],
    [193, "Espolon Blanco"],
    [4, "Bacardi FB"],
    [187, "Bud Light"],
]

# --------------------------------------------------------------------------
# PMIX workbook 3: an `All levels` sheet with NO Type column.
# Rollups cannot be told from leaves, so the parser must refuse the sheet.
# --------------------------------------------------------------------------
PMIX_NO_TYPE_ALL_LEVELS = [
    ["Menu", "Menu group", "Item, open item", "Qty sold"],
    ["LIQUOR", "Tequila", "", 193],
    ["LIQUOR", "Tequila", "Espolon Blanco", 193],
    ["BOTTLED BEER", "Domestics", "Bud Light", 187],
]

ITEM_SELECTION_HEADER = [
    "Location", "Order #", "Sent Date", "Menu Item", "Menu Group", "Menu",
    "Sales Category", "Net Price", "Qty", "Void?",
]

# One business day (2026-08-30, Sunday) at the Short North, plus decoys.
ITEM_SELECTION_ROWS = [
    # Same item, different menus and categories -- the case that rules out
    # name-only and category-only matching.
    ["Townhall - Short North", "17", "8/30/2026 19:06", "Espolon Blanco", "Tequila", "LIQUOR", "Liquor", "9.00", "2", "FALSE"],
    ["Townhall - Short North", "18", "8/30/2026 19:40", "Espolon Blanco", "Tequila Bottles", "BOTTLE SERVICE", "Bottle Service", "220.00", "1", "FALSE"],
    # Blank Sales Category on a genuine liquor SKU.
    ["Townhall - Short North", "19", "8/30/2026 20:15", "Bacardi FB", "Rum", "LIQUOR", "", "8.00", "3", "FALSE"],
    ["Townhall - Short North", "20", "8/30/2026 21:00", "Bud Light", "Domestics", "BOTTLED BEER", "Bottled Beer", "5.00", "12", "FALSE"],
    ["Townhall - Short North", "21", "8/30/2026 21:30", "Nutrl", "Seltzer", "BOTTLED BEER", "Bottled Beer", "6.00", "4", "FALSE"],
    # A void: excluded from depletion entirely.
    ["Townhall - Short North", "22", "8/30/2026 22:00", "Bud Light", "Domestics", "BOTTLED BEER", "Bottled Beer", "5.00", "6", "TRUE"],
    # After midnight but before 4 AM: still Sunday's business day.
    ["Townhall - Short North", "23", "8/31/2026 1:30", "Bud Light", "Domestics", "BOTTLED BEER", "Bottled Beer", "5.00", "5", "FALSE"],
    # 4:05 AM: the new business day has started.
    ["Townhall - Short North", "24", "8/31/2026 4:05", "Bud Light", "Domestics", "BOTTLED BEER", "Bottled Beer", "5.00", "1", "FALSE"],
    # Another location in the same Toast account -- must be filtered out.
    ["TH Ohio City", "25", "8/30/2026 19:10", "Bud Light", "Domestics", "BOTTLED BEER", "Bottled Beer", "5.00", "99", "FALSE"],
    # An item with no mapping: goes to the unmapped queue, never dropped.
    ["Townhall - Short North", "26", "8/30/2026 22:30", "Mystery Shot", "Shots", "LIQUOR", "Liquor", "7.00", "8", "FALSE"],
    # Trailing totals row: every field blank except Qty.
    ["", "", "", "", "", "", "", "", "36", ""],
]

# Variant CSV: different column ORDER, plus Comp? and Modifiers columns that the
# observed export does not have. Exercises named matching, comp inclusion and
# pour modifiers in one file.
ITEM_SELECTION_MODS_HEADER = [
    "Order #", "Void?", "Sent Date", "Qty", "Menu Item", "Sales Category",
    "Menu Group", "Menu", "Net Price", "Comp?", "Modifiers", "Location",
]
ITEM_SELECTION_MODS_ROWS = [
    ["101", "FALSE", "8/30/2026 20:00", "1", "Espolon Blanco", "Liquor", "Tequila", "LIQUOR", "9.00", "FALSE", "Single", "Townhall - Short North"],
    ["102", "FALSE", "8/30/2026 20:05", "1", "Espolon Blanco", "Liquor", "Tequila", "LIQUOR", "12.00", "FALSE", "Double", "Townhall - Short North"],
    ["103", "FALSE", "8/30/2026 20:10", "1", "Espolon Blanco", "Liquor", "Tequila", "LIQUOR", "11.00", "FALSE", "Rocks", "Townhall - Short North"],
    # No modifier at all: must default to Single.
    ["104", "FALSE", "8/30/2026 20:15", "1", "Espolon Blanco", "Liquor", "Tequila", "LIQUOR", "9.00", "FALSE", "", "Townhall - Short North"],
    # Comped: counted as depletion, flagged separately.
    ["105", "FALSE", "8/30/2026 20:20", "2", "Bud Light", "Bottled Beer", "Domestics", "BOTTLED BEER", "0.00", "TRUE", "", "Townhall - Short North"],
    # Voided: excluded.
    ["106", "TRUE", "8/30/2026 20:25", "4", "Bud Light", "Bottled Beer", "Domestics", "BOTTLED BEER", "5.00", "FALSE", "", "Townhall - Short North"],
    # An unclassified modifier: goes to the unmapped queue with its quantity.
    ["107", "FALSE", "8/30/2026 20:30", "1", "Espolon Blanco", "Liquor", "Tequila", "LIQUOR", "9.00", "FALSE", "Extra Dirty", "Townhall - Short North"],
]

#: Weekday demand shape used to synthesize five weeks of history.
HISTORY_WEEKDAY_UNITS = {
    0: {"Bud Light": 40, "Nutrl": 8, "Espolon Blanco": 10, "Red Bull": 6, "House IPA Pint": 30},
    1: {"Bud Light": 38, "Nutrl": 7, "Espolon Blanco": 9, "Red Bull": 5, "House IPA Pint": 28},
    2: {"Bud Light": 45, "Nutrl": 9, "Espolon Blanco": 12, "Red Bull": 7, "House IPA Pint": 33},
    3: {"Bud Light": 55, "Nutrl": 12, "Espolon Blanco": 15, "Red Bull": 9, "House IPA Pint": 40},
    4: {"Bud Light": 90, "Nutrl": 20, "Espolon Blanco": 26, "Red Bull": 15, "House IPA Pint": 65},
    5: {"Bud Light": 110, "Nutrl": 26, "Espolon Blanco": 32, "Red Bull": 18, "House IPA Pint": 80},
    6: {"Bud Light": 60, "Nutrl": 14, "Espolon Blanco": 17, "Red Bull": 10, "House IPA Pint": 45},
}

ITEM_META = {
    "Bud Light": ("Domestics", "BOTTLED BEER", "Bottled Beer", "5.00"),
    "Nutrl": ("Seltzer", "BOTTLED BEER", "Bottled Beer", "6.00"),
    "Espolon Blanco": ("Tequila", "LIQUOR", "Liquor", "9.00"),
    "Red Bull": ("NA", "NA BEVERAGE", "NA Beverage", "4.00"),
    "House IPA Pint": ("Draft", "DRAFT", "Bottled Beer", "7.00"),
}

#: Share of a business day's sales landing in each hour after the 4 AM cutoff.
#: Slot 0 = 04:00. The bar sells nothing until late afternoon.
HOURLY_SHAPE = {
    12: 0.02,  # 16:00
    13: 0.04,  # 17:00
    14: 0.06,
    15: 0.10,
    16: 0.14,  # 20:00
    17: 0.16,
    18: 0.16,
    19: 0.14,
    20: 0.10,  # 24:00
    21: 0.05,
    22: 0.02,  # 02:00
    23: 0.01,
}


def _write_sheet(workbook: Workbook, title: str, rows: list[list]) -> None:
    sheet = workbook.create_sheet(title)
    for row in rows:
        sheet.append(row)


def build_pmix_full(path: Path) -> None:
    """PMIX with every sheet, a Sales Category column, and rollup rows."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_sheet(workbook, "Summary", [["Report", "Product Mix"], ["Range", "8/24/2026 - 8/30/2026"]])
    _write_sheet(workbook, "All levels", PMIX_FULL_ALL_LEVELS)
    _write_sheet(workbook, "Menus", [["Menu", "Qty sold"], ["LIQUOR", 400]])
    _write_sheet(workbook, "Menu groups", [["Menu group", "Qty sold"], ["Tequila", 193]])
    _write_sheet(workbook, "Items", PMIX_FULL_ITEMS)
    _write_sheet(workbook, "Open items", PMIX_FULL_OPEN_ITEMS)
    _write_sheet(workbook, "Modifiers", [])          # present but EMPTY, as observed
    _write_sheet(workbook, "Special requests", [])
    workbook.save(path)


def build_pmix_variant(path: Path) -> None:
    """PMIX from the same account with a different sheet and column set."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_sheet(workbook, "Summary", [["Report", "Product Mix"]])
    _write_sheet(workbook, "All levels", PMIX_VARIANT_ALL_LEVELS)
    _write_sheet(workbook, "Items", PMIX_VARIANT_ITEMS)
    workbook.save(path)


def build_pmix_no_type(path: Path) -> None:
    """PMIX whose All levels sheet lost its Type column at export time."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_sheet(workbook, "All levels", PMIX_NO_TYPE_ALL_LEVELS)
    _write_sheet(workbook, "Items", PMIX_FULL_ITEMS)
    workbook.save(path)


def build_item_selection(path: Path) -> None:
    """ItemSelectionDetails with the observed header, verbatim."""
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(ITEM_SELECTION_HEADER)
        writer.writerows(ITEM_SELECTION_ROWS)


def build_item_selection_mods(path: Path) -> None:
    """ItemSelectionDetails variant: shuffled columns, comps and modifiers."""
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(ITEM_SELECTION_MODS_HEADER)
        writer.writerows(ITEM_SELECTION_MODS_ROWS)


def build_history(path: Path, *, weeks: int = 5, end: str = "2026-08-30") -> None:
    """Five weeks of timestamped line items, ending Sunday 2026-08-30.

    Quantities follow a fixed weekday shape with a small deterministic jitter, so
    trailing means and the hourly profile are both stable across runs.
    """
    rng = random.Random(20260830)
    last_day = datetime.strptime(end, "%Y-%m-%d").date()
    first_day = last_day - timedelta(days=weeks * 7 - 1)
    order_number = 1000
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(ITEM_SELECTION_HEADER)
        day = first_day
        while day <= last_day:
            shape = HISTORY_WEEKDAY_UNITS[day.weekday()]
            for item, base_units in shape.items():
                group, menu, category, price = ITEM_META[item]
                units = max(1, int(round(base_units * rng.uniform(0.92, 1.08))))
                for slot, share in HOURLY_SHAPE.items():
                    count = int(round(units * share))
                    if count <= 0:
                        continue
                    hour = (4 + slot) % 24
                    stamp_day = day + timedelta(days=1) if hour < 4 else day
                    stamp = datetime.combine(stamp_day, datetime.min.time()).replace(
                        hour=hour, minute=15
                    )
                    order_number += 1
                    writer.writerow([
                        "Townhall - Short North",
                        str(order_number),
                        stamp.strftime("%-m/%-d/%Y %-H:%M"),
                        item, group, menu, category, price, str(count), "FALSE",
                    ])
            day += timedelta(days=1)


def build_all() -> list[Path]:
    """Write every fixture. Returns the paths written."""
    written = []
    for name, builder in (
        ("pmix_full.xlsx", build_pmix_full),
        ("pmix_variant_subgroup_no_category.xlsx", build_pmix_variant),
        ("pmix_all_levels_no_type.xlsx", build_pmix_no_type),
        ("item_selection_details.csv", build_item_selection),
        ("item_selection_details_with_mods.csv", build_item_selection_mods),
        ("item_selection_history.csv", build_history),
    ):
        path = HERE / name
        builder(path)
        written.append(path)
    return written


if __name__ == "__main__":
    for path in build_all():
        print(f"wrote {path}")
