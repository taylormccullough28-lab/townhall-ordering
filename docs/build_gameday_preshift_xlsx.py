import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import column_index_from_string

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Game Day Pre-Shift"

F = lambda s=12, b=False, i=False, c="000000": Font(name="Calibri", size=s, bold=b, italic=i, color=c)
B12, R12, B14, B11, R11 = F(12, True), F(12), F(14, True), F(11, True), F(11)
BI11 = F(11, True, True)
I10 = F(10, False, True, "595959")

med, thin = Side(style="medium"), Side(style="thin")
BOX = Border(top=med, bottom=med, left=med, right=med)
THINBOX = Border(top=thin, bottom=thin, left=thin, right=thin)
UNDER = Border(bottom=thin)
GREY = PatternFill("solid", fgColor="EDEDED")
DARK = PatternFill("solid", fgColor="D9D9D9")

CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
TOPL = Alignment(horizontal="left", vertical="top", wrap_text=True)

for col, w in {"A": 20.5, "B": 11, "C": 15, "D": 20, "E": 17, "F": 11, "G": 11}.items():
    ws.column_dimensions[col].width = w

def band(row, c1, c2, value, font=B12, align=LEFT, border=None, fill=None, height=None):
    ws.merge_cells(start_row=row, end_row=row, start_column=c1, end_column=c2)
    cell = ws.cell(row=row, column=c1)
    cell.value, cell.font, cell.alignment = value, font, align
    if border or fill:
        for cc in range(c1, c2 + 1):
            t = ws.cell(row=row, column=cc)
            if border: t.border = border
            if fill: t.fill = fill
    if height: ws.row_dimensions[row].height = height

def box(row1, row2, c1, c2, height=18):
    ws.merge_cells(start_row=row1, end_row=row2, start_column=c1, end_column=c2)
    ws.cell(row=row1, column=c1).alignment = TOPL
    for r in range(row1, row2 + 1):
        ws.row_dimensions[r].height = height
        for cc in range(c1, c2 + 1):
            ws.cell(row=r, column=cc).border = THINBOX

def lab(row, col, text, font=B12):
    c = ws.cell(row=row, column=col)
    c.value, c.font, c.alignment = text, font, LEFT

def rule(row, c1, c2):
    for c in range(c1, c2 + 1):
        ws.cell(row=row, column=c).border = UNDER

def checks(row, items):
    for item in items:
        c = ws.cell(row=row, column=1)
        c.value, c.font, c.alignment = "☐", F(14), CEN
        band(row, 2, 7, item, R11, LEFT, height=17)
        row += 1
    return row

r = 1
band(r, 1, 7, "TOWNHALL COLUMBUS  —  OSU GAME DAY PRE-SHIFT", B14, CEN, fill=DARK, height=24); r += 1
band(r, 1, 7, "Read this at pre-shift. Do not summarize it.", I10, CEN); r += 2

for left, right in [("DATE:", "MATCHUP:"), ("DAY:", "EXPECTED VOLUME:"),
                    ("SHIFT:  AM / PM", "MANAGER ON DUTY:")]:
    lab(r, 1, left);  rule(r, 2, 3)
    lab(r, 4, right); rule(r, 6, 7)
    r += 1
r += 1

band(r, 1, 7, "THE THREE WINDOWS — WRITE THE TIMES ON THE BOARD", B12, LEFT, fill=GREY, height=18); r += 1
for name, note in [
    ("PRE-GAME", "Volume and speed, turn-and-burn. Guests have a hard out."),
    ("GAME WINDOW", "Floor thins, bar fills. It feels slow. It is not — it is the setup."),
    ("POST-GAME", "Hardest 90 minutes of the week. All hands on the floor before the final whistle."),
]:
    lab(r, 1, name, B11); rule(r, 2, 2)
    band(r, 3, 7, note, R11, LEFT, height=16); r += 1
r += 1

band(r, 1, 7, "NATIONAL ANTHEM", B12, LEFT, fill=GREY, height=18); r += 1
band(r, 1, 7, "When the anthem starts, the staff lines up behind the bar and faces the projector screen. All of you.", B11, LEFT, height=17); r += 1
band(r, 1, 7, "Everything stops. No pouring, no running food, no ringing tickets, no side conversations. Hands empty, hats off, eyes on the screen. It is ninety seconds, and in a room full of people who have never been here, it is the ninety seconds they remember. Get the floor ready before it starts, not during it.", R11, LEFT, height=44); r += 1
for left, right in [("ANTHEM TIME:", "WHO CALLS THE FLOOR TO STOP:")]:
    lab(r, 1, left, R11); rule(r, 2, 2)
    lab(r, 3, right, R11); ws.merge_cells(start_row=r, end_row=r, start_column=3, end_column=5)
    rule(r, 6, 7); ws.row_dimensions[r].height = 18
    r += 1
r += 1

band(r, 1, 7, "THE THREE HARD LINES — SAY ALL THREE OUT LOUD, EVERY GAME DAY", B12, CEN, border=BOX, fill=DARK, height=20); r += 1
for i, line in enumerate([
    '1.  NO DRINKING — not before your shift, not on your shift, not on the clock in any capacity. Not a taste, not a guest-bought round, not overtime.',
    '2.  NO PROMO BOTTLES — no bottle leaves the bar as a promo, comp, gift or bucket. Every bottle rings in and gets paid for. Every single one.',
    '3.  NO DISCOUNTS — every check rings at full price. No employee rate for a friend, no "take care of them." You do not have discount authority.',
    "NOTE: a promo bottle is NOT bottle service. Bottle service is a product we SELL and ring at full price — sell it hard. A promo bottle is one that leaves with no ticket. Sell every bottle. Give away none.",
    "If a guest needs to be taken care of, GET A MANAGER. You will never get in trouble for getting a manager. You will get in trouble for making the call yourself.",
]):
    band(r, 1, 7, line, B11 if i < 3 else BI11, LEFT, border=BOX, height=30); r += 1  # 4th+ lines italic
r += 1

band(r, 1, 7, "ID'ING — NO EXCEPTIONS, NOT ON GAME DAY", B12, CEN, border=BOX, fill=DARK, height=20); r += 1
for line in [
    "Every guest who looks under 40 gets carded. Every time. At the bar, at the table, and again if someone hands them a drink they did not order.",
    "Physical ID only. Expired is not an ID. A vertical license means look twice and do the math out loud.",
    "If you are not sure, you do not pour. Hand it to a manager — that is the whole process, and nobody gets second-guessed for using it.",
    "Game day is the night we get tested on this. The liquor license is the entire business. One bad pour costs more than the whole night made.",
]:
    band(r, 1, 7, line, R11, LEFT, border=BOX, height=28); r += 1
r += 1

band(r, 1, 7, "THE BOARD", B12, LEFT, fill=GREY, height=18); r += 1
lab(r, 1, "GAME TIME:", B11); rule(r, 2, 3)
lab(r, 4, "DOORS:", B11); rule(r, 6, 7)
ws.row_dimensions[r].height = 19
r += 2

lab(r, 1, "GOAL / CONTEST FOR THE DAY:"); r += 1
box(r, r + 1, 1, 7, 20); r += 3

lab(r, 1, "NEW BEERS / NEW ITEMS:"); r += 1
box(r, r + 1, 1, 7); r += 3

lab(r, 1, "86:"); r += 1
box(r, r + 1, 1, 7); r += 3

lab(r, 1, "OTHER IMPORTANT NOTES:"); r += 1
box(r, r + 1, 1, 7); r += 3

band(r, 1, 7, "PROMOS FOR THE DAY — EVERYBODY KNOWS ALL THREE BEFORE DOORS", B12, LEFT, fill=GREY, height=18); r += 1
for name, price, note in [
    ("Suncruiser Buckets", "6 / 30", "Lead with this at every table in the pre-game window. Fastest ticket in the building."),
    ("Lucky One Lemonade", "6 / 24", "Same play. Buckets land on the table before the first food order goes in."),
    ("Tito's Bottle Service", "PUSH HARD", "The one we are chasing tonight. Offer it to every large party, every booth, every group already celebrating. Rings at full price — bottle service is sold, never comped."),
]:
    lab(r, 1, name, B11)
    lab(r, 3, price, F(11, True, False, "A4211B" if price == "PUSH HARD" else "2C574A"))
    band(r, 4, 7, note, R11, LEFT, height=30)
    r += 1
lab(r, 1, "Other promos:", R11); rule(r, 2, 7); ws.row_dimensions[r].height = 18
r += 2

# ---- staff counts & completions, AM and PM ----
STAFF_COUNTS = ["SERVERS:", "BARTENDERS:", "CAFÉ:", "HOSTS:", "RUNNERS:", "BARBACKS:", "SHOT GIRL:", "MANAGERS:"]
SHIFT_LEADS  = ["SERVER:", "CAFÉ:", "BARTENDER:", "HOST:", "BAR:", "BARBACK:"]
COMPLETIONS  = ["SERVER:", "BAR:", "CAFÉ:", "HOST:", "BARBACK:"]

def stack_group(top, col, title, items):
    """Bold group title with a named, ruled line per row. Returns the row after the last item."""
    lab(top, col, title, B11)
    for i, s in enumerate(items):
        lab(top + 1 + i, col, s, R12)
        rule(top + 1 + i, col + 1, 7 if col == 5 else col + 1)
        ws.row_dimensions[top + 1 + i].height = 18
    return top + 1 + len(items)

for daypart in ("AM", "PM"):
    band(r, 1, 7, daypart + " CREW — STAFF #'S & COMPLETIONS", B12, LEFT, fill=GREY, height=18)
    r += 1
    top = r
    # staff counts carry two write lines: how many are on, and how many are training
    lab(top, 1, "STAFF #'S", B11)
    lab(top, 2, "ON", B11)
    lab(top, 3, "IN TRAINING", B11)
    lab(top, 5, "SHIFT LEADS", B11)
    for i, role in enumerate(STAFF_COUNTS):
        rr = top + 1 + i
        lab(rr, 1, role, R12)
        rule(rr, 2, 2)
        rule(rr, 3, 4)
        ws.row_dimensions[rr].height = 18
    for i, s in enumerate(SHIFT_LEADS):
        rr = top + 1 + i
        lab(rr, 5, s, R12)
        rule(rr, 6, 7)
    first, last = top + 1, top + len(STAFF_COUNTS)
    tr = max(last, top + len(SHIFT_LEADS)) + 1
    lab(tr, 1, "TOTAL SCHEDULED:", B11)
    for col, letter in ((2, "B"), (3, "C")):
        cell = ws.cell(row=tr, column=col)
        cell.value = "=SUM({0}{1}:{0}{2})".format(letter, first, last)
        cell.font, cell.alignment = B12, LEFT
        cell.border = Border(top=thin)
    ws.cell(row=tr, column=4).border = Border(top=thin)
    ws.row_dimensions[tr].height = 19

    tr2 = tr + 1
    lab(tr2, 1, "TOTAL STAFF ON SHIFT:", B11)
    total = ws.cell(row=tr2, column=2)
    total.value = "=B{0}+C{0}".format(tr)
    total.font, total.alignment = B12, LEFT
    lab(tr2, 3, "(scheduled + in training)", F(10, False, True, "595959"))
    ws.row_dimensions[tr2].height = 19

    r = tr2 + 2
    ends = [stack_group(r, 1, "STAFF TO COACH — 1 HUDDL", COMPLETIONS),
            stack_group(r, 3, "7SHIFTS COMPLETION", COMPLETIONS)]
    r = max(ends) + 1

band(r, 1, 7, "STATION ASSIGNMENTS — WRITE A NAME ON EVERY LINE", B12, LEFT, fill=GREY, height=18)
r += 1

STATIONS = [
    ("HOST", ["First in command:", "Second in command:", "Runner:", "Waters:", "Bathrooms:"]),
    ("RUNNERS", ["Bar 100:", "Bar 200:", "Food:"]),
    ("BAR", ["100:", "105:", "110:", "200:", "Patio:"]),
    ("BARISTA", ["POS 1:", "POS 2:", "Cocktail 1:", "Cocktail 2:"]),
    ("BARBACKS", ["Main Bar:", "Patio Bar:", "Main Floor:", "Café:", "Front Patio:"]),
    ("SHOT GIRL", ["On tonight:"]),
    ("TASTE PLATE", ["Running it:"]),
    ("SOCIAL MEDIA", ["On tonight:", "Backup / 2nd half:"]),
]

for daypart in ("AM", "PM"):
    band(r, 1, 7, daypart + " CREW", B11, LEFT, height=17)
    r += 1
    for i in range(0, len(STATIONS), 3):     # three groups fit across columns A, C and E
        top = r
        ends = [stack_group(top, 1 + 2 * j, title, items)
                for j, (title, items) in enumerate(STATIONS[i:i + 3])]
        r = max(ends) + 1
    r += 1

lab(r, 1, "PARTIES / EVENTS:"); r += 1
box(r, r + 1, 1, 7); r += 3

band(r, 1, 7, "GAME DAY OPEN — 90 MINUTES BEFORE DOORS", B12, LEFT, fill=GREY, height=18); r += 1
r = checks(r, [
    "Know the game — kickoff, attendance, national TV, weather. Times on the board.",
    "Every staff member can name all three promos — ask two people at random before pre-shift ends.",
    "Anthem time on the board — every staff member knows to line up behind the bar.",
    "Social media assigned, phone charged, and they know the shot list.",
    "Three windows built on paper — cut times, breaks, post-game all-hands decided now.",
    "Staffing confirmed against 7shifts. One no-show on game day is a two-hour wait.",
    "Every station walked, front and back. The 3 Rules on all of them.",
    "Bar par DOUBLED — ice, kegs, backups, garnish, glassware, batch.",
    "86 board walked with the kitchen, current and visible.",
    "TVs, sound and the game feed tested before a single guest sits down.",
    "ID check covered — door and bar know who is carding, and the light works.",
    "Restrooms stocked and clean before doors.",
    "POS, printers and card processing tested with a live ticket.",
    "Cash on hand for a bar that will run cash all night.",
    "Door, patio and line plan — where the wait forms and who holds it.",
])
r += 1

band(r, 1, 7, "UNIFORM CHECK", B12, LEFT, fill=GREY, height=18)
r += 1
r = checks(r, ["OSU body suit and / or polo", "Eye black", "Glitter", "Ribbon", "Hair pulled back - tight", "Makeup done", "Hair done", "Gentlemen have a belt on"])
r += 1

for l, rt in [("SERVERS HAVE WINE KEY & LIGHTER?   YES / NO", "ID'ING EVERYONE UNDER 40?   YES / NO"),
              ("TVs & GAME FEED TESTED?   YES / NO", "BAR PAR DOUBLED FOR POST-GAME?   YES / NO")]:
    band(r, 1, 3, l, B12, CEN, border=BOX, height=20)
    band(r, 4, 7, rt, B12, CEN, border=BOX)
    r += 1
band(r, 1, 7, "**REMIND STAFF TO TEXT FRIENDS TO COME VISIT**", B12, CEN, border=BOX, height=20); r += 2

band(r, 1, 7, "BY POSITION — READ THE BLOCK THAT BELONGS TO EACH GROUP", B12, LEFT, fill=GREY, height=18); r += 1
for name, note in [
    ("BAR", "ID before you pour — every time, no exceptions. Every drink rings before it pours. No promo bottles, no comps. Cut people early and tell a manager immediately. Stock backups during the game window, not after it."),
    ("SERVERS", 'Greet inside 30 seconds. Ask "are you trying to be out by kickoff?" first, then fire accordingly. Card every table before the first round lands. Full hands in, full hands out. Guest issue goes to a manager immediately.'),
    ("HOST", "Own the door and own the wait. An accurate quote beats a short one — give a real number, then beat it. First in command runs the book; second in command steps in the moment first gets pulled. Card at the door when the room is young. Keep the line off the sidewalk and feed the wait to the bar."),
    ("KITCHEN", "Prep to game-day pars, confirmed before service. 86 goes to the manager when you see it coming, not when you hit zero. Ticket times called every window. Original position, original condition at every changeover."),
    ("SOCIAL", "Work the room all night, not just the first hour. Get the staff lined up behind the bar during the anthem, buckets landing on tables, the bottle service pour, and the bar at full tilt. Shoot people who want to be shot — ask first, and never a minor, a check, or a POS screen. Post in the game window, not in the crush."),
    ("SUPPORT", "You set the pace of the building. Tables turning is the only thing standing between us and a two-hour wait. Everybody runs food."),
]:
    lab(r, 1, name, B11)
    band(r, 2, 7, note, R11, LEFT, height=32); r += 1
r += 1

band(r, 1, 7, "POST-GAME & CLOSE", B12, LEFT, fill=GREY, height=18); r += 1
r = checks(r, [
    "All hands on the floor BEFORE the final whistle — not after.",
    "Full bottle count and bar inventory tonight, not tomorrow.",
    "Every void, comp and discount reconciled to a manager name.",
    "Every refusal or cut-off logged with what happened and who made the call.",
    "Every station back to original position, original condition.",
    "Shift notes written: what we ran out of, where the wait broke down, what changes next game.",
])
r += 1

lab(r, 1, "MARKETING:"); r += 1
box(r, r + 1, 1, 7); r += 3


band(r, 1, 7, "100% of our standards, 100% of the time, at 100% volume. A full house is not an excuse to run at 80% — it is the reason we hold the line.",
     BI11, CEN, height=28)

ws.page_setup.orientation = "portrait"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
for m in ("left", "right", "top", "bottom"):
    setattr(ws.page_margins, m, 0.4)
ws.print_area = f"A1:G{r}"

out = "/home/user/townhall-ordering/docs/GameDay_PreShift_Template.xlsx"
wb.save(out)
print("saved", out, "last row", r)
