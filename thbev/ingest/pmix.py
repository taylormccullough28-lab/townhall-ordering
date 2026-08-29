"""Parser for Toast PMIX multi-sheet XLSX exports.

Structure confirmed against two recovered exports (PRD ingest appendix). Up to
eight sheets appear -- ``Summary``, ``All levels``, ``Menus``, ``Menu groups``,
``Items``, ``Open items``, ``Modifiers``, ``Special requests`` -- and *both the
sheet set and the column set vary between exports from the same account*,
because they are selected at export time.

Three rules this parser exists to enforce:

1. Columns are matched by normalized header name, never by position.
2. ``All levels`` mixes rollup rows (blank ``Type``) with leaf rows
   (``menuItem`` / ``openItem``). Only leaves are real sales; counting rollups
   double-counts every quantity in the file.
3. ``Sales Category`` is optional and is blank on genuine liquor SKUs, so it can
   never be the sole beverage filter. The parser preserves blanks rather than
   dropping the row.

PMIX carries no time dimension. Rows come back with ``sold_at`` and
``business_date`` set to None, which is what makes the post-cutoff adjustment
refuse to run on PMIX alone.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Sequence

from ..models import IngestResult, ItemKey, SalesRow, Severity, UnmappedRow
from ..normalize import cell, clean_text, normalize_header, parse_number, squash
from .columns import ALL_LEVELS_ALIASES, ITEMS_ALIASES, LEAF_TYPES, MODIFIER_ALIASES, SHEET_ROLES
from .sheets import find_header_row, is_blank_row, looks_like_totals_row

EXPECTED_SHEETS = tuple(SHEET_ROLES)


def parse_pmix(
    path: str | Path,
    *,
    prefer: str = "all_levels",
    include_modifiers: bool = True,
) -> IngestResult:
    """Parse a PMIX workbook into normalized sales rows.

    Args:
        path: Path to the ``.xlsx`` export.
        prefer: ``"all_levels"`` (default) uses the sheet that carries menu
            structure, which the composite mapping key needs, and falls back to
            ``Items``/``Open items`` when it is absent or unusable.
            ``"items"`` forces the flat sheets.
        include_modifiers: Parse the ``Modifiers`` sheet when present. Every
            recovered export either omits it or leaves it empty; rows found
            there are returned as modifier-tagged sales rows for the caller to
            classify.

    Returns:
        An :class:`~thbev.models.IngestResult`. Rows that survive filtering but
        cannot be normalized land in ``unmapped`` with their quantities intact.

    Raises:
        FileNotFoundError: If the workbook does not exist.
        ValueError: If the file is not a readable workbook.
    """
    from openpyxl import load_workbook

    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"PMIX workbook not found: {path}")

    result = IngestResult(source_files=[str(path)])
    try:
        workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises a variety of types
        raise ValueError(f"Could not read {path} as an XLSX workbook: {exc}") from exc

    try:
        sheets_by_role: dict[str, list[str]] = {}
        for name in workbook.sheetnames:
            result.sheets_present.append(name)
            role = SHEET_ROLES.get(normalize_header(name))
            if role:
                sheets_by_role.setdefault(role, []).append(name)

        for expected in EXPECTED_SHEETS:
            role = SHEET_ROLES[expected]
            if role not in sheets_by_role:
                result.add_issue(
                    Severity.INFO,
                    "sheet_absent",
                    f"Sheet '{expected}' is not in this export; every sheet is optional.",
                    sheet=expected,
                )

        used_all_levels = False
        if prefer == "all_levels" and "all_levels" in sheets_by_role:
            for name in sheets_by_role["all_levels"]:
                rows = _read_sheet(workbook[name])
                used_all_levels |= _parse_all_levels(name, rows, result)

        if not used_all_levels:
            if prefer == "all_levels":
                result.add_issue(
                    Severity.WARNING,
                    "all_levels_unavailable",
                    "Falling back to the flat Items/Open items sheets; menu structure "
                    "is unavailable so the composite mapping key will match on "
                    "category + item only.",
                )
            for role in ("items", "open_items"):
                for name in sheets_by_role.get(role, []):
                    rows = _read_sheet(workbook[name])
                    _parse_items(name, rows, result, open_items=(role == "open_items"))

        if include_modifiers:
            for name in sheets_by_role.get("modifiers", []):
                rows = _read_sheet(workbook[name])
                _parse_modifiers(name, rows, result)
    finally:
        workbook.close()

    if not result.rows:
        result.add_issue(
            Severity.ERROR,
            "no_rows",
            "No usable sales rows were parsed from this workbook.",
        )
    return result


def _read_sheet(sheet: Any) -> list[tuple[Any, ...]]:
    return [tuple(row) for row in sheet.iter_rows(values_only=True)]


def _parse_all_levels(
    sheet_name: str, rows: Sequence[Sequence[Any]], result: IngestResult
) -> bool:
    """Parse the ``All levels`` sheet. Returns True if it produced usable rows."""
    header = find_header_row(rows, ALL_LEVELS_ALIASES)
    if header is None:
        result.add_issue(
            Severity.ERROR,
            "header_not_found",
            f"Could not find a header row on sheet '{sheet_name}'.",
            sheet=sheet_name,
        )
        return False

    header_index, columns, missing, unrecognized = header
    result.missing_columns[sheet_name] = missing
    if unrecognized:
        result.add_issue(
            Severity.INFO,
            "unrecognized_columns",
            f"Sheet '{sheet_name}' has columns this parser does not use: {', '.join(unrecognized)}.",
            sheet=sheet_name,
            columns=unrecognized,
        )

    if "type" not in columns:
        # Without Type we cannot tell rollups from leaves, and counting the
        # sheet anyway would double-count every quantity. Refuse the sheet
        # loudly rather than silently returning inflated numbers.
        body = [r for r in rows[header_index + 1 :] if not is_blank_row(r)]
        result.add_issue(
            Severity.ERROR,
            "all_levels_no_type_column",
            f"Sheet '{sheet_name}' has no 'Type' column, so rollup rows cannot be "
            f"distinguished from leaf rows. Skipped {len(body)} rows rather than "
            "double-count them. Re-export with the Type column, or use the Items sheet.",
            sheet=sheet_name,
            rows_skipped=len(body),
        )
        result.bump("all_levels_rows_skipped_no_type", len(body))
        return False

    if "qty" not in columns:
        result.add_issue(
            Severity.ERROR,
            "missing_qty_column",
            f"Sheet '{sheet_name}' has no quantity column; nothing can be counted from it.",
            sheet=sheet_name,
        )
        return False

    produced = 0
    for offset, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if is_blank_row(row):
            continue
        raw = _raw_dict(row, columns)

        if looks_like_totals_row(row, columns, ("type", "menu", "menu_item")):
            result.bump("totals_rows_dropped")
            continue

        type_value = squash(cell(row, columns, "type"))
        if type_value not in LEAF_TYPES:
            # Blank Type is a rollup subtotal; any other value is a level we do
            # not count. Both are tracked, neither is counted.
            result.bump("rollup_rows_skipped")
            result.bump(f"rollup_type:{type_value or 'blank'}")
            continue

        item = clean_text(cell(row, columns, "menu_item"))
        qty = parse_number(cell(row, columns, "qty"))
        key = ItemKey(
            sales_category=clean_text(cell(row, columns, "sales_category")),
            menu=clean_text(cell(row, columns, "menu")),
            menu_group=_menu_group_with_subgroup(row, columns),
            menu_item=item,
        )
        if item is None or qty is None:
            result.unmapped.append(
                UnmappedRow(
                    reason="blank item name" if item is None else "unreadable quantity",
                    qty=qty,
                    source=f"pmix:{sheet_name}",
                    row_number=offset,
                    key=key,
                    raw=raw,
                )
            )
            continue

        result.rows.append(
            SalesRow(
                key=key,
                qty=qty,
                source=f"pmix:{sheet_name}",
                row_number=offset,
                raw=raw,
            )
        )
        produced += 1

    if produced:
        result.sheets_used.append(sheet_name)
    return produced > 0


def _menu_group_with_subgroup(row: Sequence[Any], columns: dict[str, int]) -> str | None:
    """Return the menu group, appending Subgroup when that optional column exists.

    One recovered export carries ``Subgroup`` and the other does not. Keeping it
    inside the group slot preserves the extra specificity without changing the
    shape of the four-part key.
    """
    group = clean_text(cell(row, columns, "menu_group"))
    subgroup = clean_text(cell(row, columns, "subgroup"))
    if group and subgroup:
        return f"{group} / {subgroup}"
    return group or subgroup


def _parse_items(
    sheet_name: str,
    rows: Sequence[Sequence[Any]],
    result: IngestResult,
    *,
    open_items: bool,
) -> None:
    """Parse a flat ``Items`` or ``Open items`` sheet: Item | Sales Category | Qty sold."""
    header = find_header_row(rows, ITEMS_ALIASES)
    if header is None:
        result.add_issue(
            Severity.ERROR,
            "header_not_found",
            f"Could not find a header row on sheet '{sheet_name}'.",
            sheet=sheet_name,
        )
        return

    header_index, columns, missing, unrecognized = header
    result.missing_columns[sheet_name] = missing
    if "sales_category" in missing:
        result.add_issue(
            Severity.WARNING,
            "no_sales_category_column",
            f"Sheet '{sheet_name}' has no Sales Category column. Classification must "
            "fall back entirely to the product mapping on menu + item.",
            sheet=sheet_name,
        )
    if "qty" not in columns:
        result.add_issue(
            Severity.ERROR,
            "missing_qty_column",
            f"Sheet '{sheet_name}' has no quantity column; nothing can be counted from it.",
            sheet=sheet_name,
        )
        return
    if unrecognized:
        result.add_issue(
            Severity.INFO,
            "unrecognized_columns",
            f"Sheet '{sheet_name}' has columns this parser does not use: {', '.join(unrecognized)}.",
            sheet=sheet_name,
            columns=unrecognized,
        )

    produced = 0
    for offset, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if is_blank_row(row):
            continue
        if looks_like_totals_row(row, columns, ("menu_item",)):
            result.bump("totals_rows_dropped")
            continue
        raw = _raw_dict(row, columns)
        item = clean_text(cell(row, columns, "menu_item"))
        qty = parse_number(cell(row, columns, "qty"))
        key = ItemKey(
            sales_category=clean_text(cell(row, columns, "sales_category")),
            menu="Open items" if open_items else clean_text(cell(row, columns, "menu")),
            menu_group=clean_text(cell(row, columns, "menu_group")),
            menu_item=item,
        )
        if item is None or qty is None:
            result.unmapped.append(
                UnmappedRow(
                    reason="blank item name" if item is None else "unreadable quantity",
                    qty=qty,
                    source=f"pmix:{sheet_name}",
                    row_number=offset,
                    key=key,
                    raw=raw,
                )
            )
            continue
        result.rows.append(
            SalesRow(
                key=key,
                qty=qty,
                source=f"pmix:{sheet_name}",
                row_number=offset,
                raw=raw,
            )
        )
        produced += 1

    if produced:
        result.sheets_used.append(sheet_name)


def _parse_modifiers(
    sheet_name: str, rows: Sequence[Sequence[Any]], result: IngestResult
) -> None:
    """Parse the ``Modifiers`` sheet if it carries anything.

    Every recovered export leaves this sheet empty or omits it, so this path is
    unexercised by real data. Modifier rows are returned tagged with
    ``modifiers=("<name>",)`` and zero product mapping; classifying them as
    pour-size versus product is the catalog's job, and anything unclassified
    ends up in the unmapped queue there.
    """
    header = find_header_row(rows, MODIFIER_ALIASES)
    if header is None:
        result.add_issue(
            Severity.INFO,
            "modifiers_sheet_empty",
            f"Sheet '{sheet_name}' has no header row -- it is present but empty. "
            "No modifier-level data exists in the account yet; every spirit pour "
            "will default to Single (1.5 oz) until a modifier export is pulled.",
            sheet=sheet_name,
        )
        return

    header_index, columns, missing, _ = header
    result.missing_columns[sheet_name] = missing
    count = 0
    for offset, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if is_blank_row(row):
            continue
        raw = _raw_dict(row, columns)
        name = clean_text(cell(row, columns, "menu_item"))
        qty = parse_number(cell(row, columns, "qty"))
        if name is None or qty is None:
            result.unmapped.append(
                UnmappedRow(
                    reason="unreadable modifier row",
                    qty=qty,
                    source=f"pmix:{sheet_name}",
                    row_number=offset,
                    raw=raw,
                )
            )
            continue
        result.rows.append(
            SalesRow(
                key=ItemKey(
                    sales_category=clean_text(cell(row, columns, "sales_category")),
                    menu=clean_text(cell(row, columns, "menu")),
                    menu_group=clean_text(cell(row, columns, "menu_group")),
                    menu_item=name,
                ),
                qty=qty,
                source=f"pmix:{sheet_name}",
                modifiers=(name,),
                row_number=offset,
                raw=raw,
            )
        )
        count += 1
    if count:
        result.sheets_used.append(sheet_name)
        result.bump("modifier_rows", count)


def _raw_dict(row: Sequence[Any], columns: dict[str, int]) -> dict[str, Any]:
    return {field: cell(row, columns, field) for field in columns}
